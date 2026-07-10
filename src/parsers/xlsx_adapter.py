import openpyxl
from datetime import date
from src.schema import ProjectPlan, ProjectSummary, TaskRecord, CommentRecord
from src.parsers.base_adapter import BaseAdapter

class XlsxAdapter(BaseAdapter):
    def parse(self, file_path: str, project_id: str) -> ProjectPlan:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        parse_warnings = []
        
        # 1. Identify sheets
        summary_sheet = None
        comments_sheet = None
        main_sheet = None
        
        for sheet_name in wb.sheetnames:
            if "summary" in sheet_name.lower():
                summary_sheet = wb[sheet_name]
            elif "comment" in sheet_name.lower():
                comments_sheet = wb[sheet_name]
            else:
                sheet = wb[sheet_name]
                if sheet.max_row > 50 and main_sheet is None:
                    main_sheet = sheet
                    
        if not main_sheet:
            main_sheet = wb.active # Fallback
            parse_warnings.append("Could not identify main task sheet definitively, using active sheet.")
            
        # 2. Parse Summary Sheet
        summary_data = {}
        if summary_sheet:
            for row in summary_sheet.iter_rows(values_only=True):
                if row[0] and isinstance(row[0], str):
                    summary_data[row[0].strip().lower()] = row[1]
                    
        project_name = summary_data.get("project name")
        if not project_name:
            project_name = file_path.split("\\")[-1].split("/")[-1].replace(".xlsx", "")
            
        summary = ProjectSummary(
            project_name=project_name,
            project_manager=summary_data.get("project manager"),
            project_start=self._parse_date(summary_data.get("project start date")),
            project_end=self._parse_date(summary_data.get("project end date")),
            pct_complete=self._parse_float(summary_data.get("% complete")),
            project_stage=summary_data.get("project stage"),
            at_risk_flag=summary_data.get("at risk"),
            source_schedule_health=summary_data.get("schedule health")
        )
        
        # 3. Parse Main Task Sheet
        tasks = []
        header_map = {}
        header_row_idx = 1
        
        for r_idx, row in enumerate(main_sheet.iter_rows(values_only=True), start=1):
            # Try to find header row (look for "Task Name" or similar)
            is_header = any(cell and isinstance(cell, str) and "task name" in cell.lower() for cell in row)
            if is_header:
                header_row_idx = r_idx
                for c_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        header_map[cell.strip().lower()] = c_idx
                break
                
        if not header_map:
            parse_warnings.append("Could not find a valid header row in main task sheet.")
            
        # Determine actual column mappings based on headers
        col_mapping = {
            "task_name": self._find_col(header_map, ["task name"]),
            "phase_milestone": self._find_col(header_map, ["phase/milestone"]),
            "level": self._find_col(header_map, ["level"]),
            "ancestors": self._find_col(header_map, ["ancestors"]),
            "status": self._find_col(header_map, ["status"]),
            "pct_complete": self._find_col(header_map, ["% complete"]),
            "baseline_start": self._find_col(header_map, ["baseline start", "baseline start date", "baseline start2"]),
            "baseline_end": self._find_col(header_map, ["baseline finish", "baseline end date", "baseline finish2"]),
            "actual_start": self._find_col(header_map, ["start date", "start"]),
            "actual_end": self._find_col(header_map, ["end date", "finish"]),
            "variance_days": self._find_col(header_map, ["variance", "variance2"]),
            "total_float": self._find_col(header_map, ["total float"]),
            "is_critical_path": self._find_col(header_map, ["critical ?"]),
            "on_hold": self._find_col(header_map, ["on hold?"]),
            "not_applicable": self._find_col(header_map, ["not applicable?"]),
            "priority": self._find_col(header_map, ["priority"]),
            "status_comment": self._find_col(header_map, ["status comment"]),
            "rag": self._find_col(header_map, ["rag", "schedule health"])
        }
        
        if col_mapping["level"] is None:
            parse_warnings.append("Level column absent, hierarchy inferred from Ancestors only or not at all.")

        for r_idx, row in enumerate(main_sheet.iter_rows(min_row=header_row_idx+1, values_only=True), start=header_row_idx+1):
            if not row or all(cell is None for cell in row):
                continue
                
            task_name_col = col_mapping["task_name"]
            task_name = str(row[task_name_col]) if task_name_col is not None and row[task_name_col] is not None else f"Row_{r_idx}"
            
            variance_val, variance_conf = self._parse_variance(row, col_mapping["variance_days"])
            
            field_conf = {}
            if variance_conf:
                field_conf["variance_days"] = variance_conf
            
            source_rag_col = col_mapping["rag"]
            if source_rag_col is not None and row[source_rag_col]:
                field_conf["source_rag"] = "direct" # Storing separate, per spec
                
            tasks.append(TaskRecord(
                task_id=f"{project_id}_{r_idx}",
                task_name=task_name,
                phase_milestone=self._get_str(row, col_mapping["phase_milestone"]),
                level=self._parse_int(self._get_val(row, col_mapping["level"])),
                ancestors=self._get_str(row, col_mapping["ancestors"]),
                status=self._get_str(row, col_mapping["status"]),
                pct_complete=self._parse_float(self._get_val(row, col_mapping["pct_complete"])),
                baseline_start=self._parse_date(self._get_val(row, col_mapping["baseline_start"])),
                baseline_end=self._parse_date(self._get_val(row, col_mapping["baseline_end"])),
                actual_start=self._parse_date(self._get_val(row, col_mapping["actual_start"])),
                actual_end=self._parse_date(self._get_val(row, col_mapping["actual_end"])),
                variance_days=variance_val,
                total_float=self._parse_int(self._get_val(row, col_mapping["total_float"])),
                is_critical_path=self._parse_bool(self._get_val(row, col_mapping["is_critical_path"])),
                on_hold=self._parse_bool(self._get_val(row, col_mapping["on_hold"])),
                not_applicable=self._parse_bool(self._get_val(row, col_mapping["not_applicable"])),
                priority=self._get_str(row, col_mapping["priority"]),
                status_comment=self._get_str(row, col_mapping["status_comment"]),
                field_confidence=field_conf
            ))
            
        # 4. Parse Comments Sheet
        comments = []
        if comments_sheet:
            for row in comments_sheet.iter_rows(values_only=True):
                # (row_ref, comment_text, author, timestamp)
                if not row or all(cell is None for cell in row):
                    continue
                    
                if len(row) >= 2 and row[1]:
                    comments.append(CommentRecord(
                        linked_row_ref=str(row[0]) if row[0] else None,
                        text=str(row[1]),
                        author=str(row[2]) if len(row) > 2 and row[2] else None,
                        timestamp=str(row[3]) if len(row) > 3 and row[3] else None
                    ))
                    
        return ProjectPlan(
            project_id=project_id,
            source_file=file_path.split("\\")[-1].split("/")[-1],
            source_format="xlsx",
            summary=summary,
            tasks=tasks,
            comments=comments,
            parse_warnings=parse_warnings
        )
        
    def _find_col(self, header_map, possibilities):
        for p in possibilities:
            if p in header_map:
                return header_map[p]
        return None
        
    def _get_val(self, row, idx):
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        return val

    def _get_str(self, row, idx):
        val = self._get_val(row, idx)
        if val is None:
            return None
        return str(val).strip()

    def _parse_date(self, val):
        if not val:
            return None
        if isinstance(val, date):
            return val
        if hasattr(val, "date"):
            return val.date()
        return None

    def _parse_float(self, val):
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None
            
    def _parse_int(self, val):
        if val is None:
            return None
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return None

    def _parse_bool(self, val):
        if not val:
            return False
        if isinstance(val, bool):
            return val
        s = str(val).strip().lower()
        return s in ("1", "true", "yes", "y", "x")
        
    def _parse_variance(self, row, idx):
        val = self._get_val(row, idx)
        if not val:
            return None, None
        
        s = str(val).strip()
        if s == "#UNPARSEABLE":
            return None, "missing"
            
        if s.endswith("d"):
            s = s[:-1]
            
        try:
            return int(float(s)), "direct"
        except ValueError:
            return None, "missing"
